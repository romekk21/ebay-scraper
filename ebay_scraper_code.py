import openpyxl
from requests import get
from bs4 import BeautifulSoup
import sys
import tkinter as tk

# Setup user input dialogue window
window = tk.Tk()
window.title('eBay Scraper')
window.geometry('400x200')
NEW_MONTH = False

def handle_yes():
    window.destroy()
    global NEW_MONTH
    NEW_MONTH = True

def handle_no():
    window.destroy()
    global NEW_MONTH
    NEW_MONTH = False
    
def ask_if_new_month():
    l = tk.Label(window, text = "New month?")
    l.config(font =("Courier", 14))
    btn_yes = tk.Button(window, text="Yes", command = handle_yes, height = 3, width = 12)
    btn_no = tk.Button(window, text="No", command = handle_no, height = 3, width = 12)

    l.pack()
    btn_yes.pack(pady=10)
    btn_no.pack()

    window.mainloop()

def parse_ebay():

    # Used in order to get rid of emojis 
    non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)

    # Used in order to get past page 5
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.0.0 Safari/537.36'}

    # Title storage
    titles = []

    # Loop 10 pages
    for page_num in range(1, 11):
        link = 'https://www.ebay.com/sch/i.html?_from=R40&_nkw=vintage+t+shirt&_sacat=1059&rt=nc&LH_Sold=1&LH_Complete=1&_pgn='+str(page_num)
        response = get(link, headers=headers, timeout = 2)
        html_soup = BeautifulSoup(response.text, 'html.parser')

        # Store all of the posts
        posts = html_soup.find_all('li', attrs={'class': 's-item'})
        
        # Get the titles
            # Translate to account for emojis
        for post in posts:
            titles.append(post.find('h3', class_='s-item__title').text.translate(non_bmp_map))

    title_data = dict()

    # Clean the data
    stopwords = ["i","ebay","shop","women","for", "/", "to", "~","a", "men’s", "men's", "&", "and","tshirt","mens", "men", "tee", "vintage", "vtg", "on", "of", "the", "t-shirt", "shirt", "shirt,", "sz", "size", "single", "-", "all", "in", "t"]

    # For each FULL title 
    for title in titles:

        # Split into a list of the words that make up the title
        words_in_title = title.split()
        
        for word in words_in_title:

            # If a word in the title is found in stopwords, remove it from the list
            if word.lower() in stopwords:
                continue

            # Important that items titled "M" are not separate from items titled "Medium"; it's the same size
            # M = Medium, L = Large, XXL = 2XL, S = Small
            if word.lower() == "m":
                word = "Medium"
            elif word.lower() == "l":
                word = "Large"
            elif word.lower() == "xxl":
                word = "2XL"
            elif word.lower() == "s":
                word = "Small"

            # Add to dictionary, keep track of how many of the same words are found
            if word.lower() in title_data:
                title_data[word.lower()] += 1
            else:
                title_data[word.lower()] = 1

    return title_data
    
def main():

    # Get the titles data of the items sold 
    title_data = parse_ebay()

    # Split the words (keys) and the counts (values) into two lists
    words_in_title = list(title_data.keys())
    count_of_words = list(title_data.values())

    # Test to see if workbook already exists
    try:
        my_wb = openpyxl.load_workbook(filename="ebay_scraper.xlsx")
        new_wb = False
    except:
        my_wb = openpyxl.Workbook()
        new_wb = True

    # Go to first sheet
    my_sheet = my_wb.worksheets[0]

    # Set value for NEW_MONTH
    ask_if_new_month()

    
    if new_wb:
        # Working in a new workbook
        for i, value in enumerate(words_in_title, 2):
            my_sheet.cell(row = i, column = 2).value = value

        for i, value in enumerate(count_of_words, 2):
            my_sheet.cell(row = i, column = 3).value = value
    else:

        # Working in an existing workbook
        
        # Find how many rows there are
        total_rows = 2
        while my_sheet.cell(row = total_rows, column = 2).value != None:
            total_rows += 1

        # Procedure that depends on new month
        if NEW_MONTH:
            start_row = total_rows
        else:
            start_row = my_sheet.cell(row = 2, column = 6).value

        # Grab the existing words and counts in the sheet
        existing_data = {}
        for i in range(start_row, total_rows):
            existing_data[my_sheet.cell(row = i, column = 2).value.lower()] = my_sheet.cell(row = i, column = 3).value

        # Update the counts
        for i in range(len(words_in_title)):
            if words_in_title[i] in existing_data:
                count_of_words[i] += existing_data[words_in_title[i]]
                
        # Overwrite the data
        row_number = start_row
        for i in range(start_row, start_row + len(words_in_title)):
            if(count_of_words[i - start_row] >= 6):
                my_sheet.cell(row = row_number, column = 2).value = words_in_title[i - start_row].capitalize()
                my_sheet.cell(row = row_number, column = 3).value = count_of_words[i - start_row]
                row_number += 1

       
        if NEW_MONTH:
            my_sheet.cell(row = 2, column = 6).value = total_rows

    my_wb.save("ebay_scraper.xlsx")

main()
